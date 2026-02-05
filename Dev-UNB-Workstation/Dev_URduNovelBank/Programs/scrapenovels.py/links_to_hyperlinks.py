import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 📌 Excel file ka naam (Aap apni file ka naam yahan likhein)
input_file = r"D:\UNB\Programs\scrapenovels.py\Novels excel files\itsurdu\Blogger_Novels.xlsx"
output_file = r"D:\UNB\Programs\scrapenovels.py\Novels excel files\itsurdu\Blogger_Novels-hyper.xlsx"

# 📌 Excel file load karein
df = pd.read_excel(input_file)

# 📌 Workbook & Sheet ko open karein
wb = load_workbook(input_file)
ws = wb.active

# 📌 Har row me check karein aur links ko Hyperlink banayein
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("http"):  
            cell.hyperlink = cell.value  # ✅ Hyperlink set karein
            cell.font = Font(color="0000FF", underline="single")  # ✅ Blue color & underline

# 📌 Save the updated file
wb.save(output_file)

print(f"✅ Successfully converted links to hyperlinks! Saved as {output_file}")
